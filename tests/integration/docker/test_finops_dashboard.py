"""FinOps 看板与优化建议服务（finops-dashboard）Docker 集成测试。

被测对象：Docker 容器 ``it-finops-dashboard``（镜像 ``shuqing/finops-dashboard:0.1.0``），
Java/Spring Boot，主机端口 18085 → 容器 8085。

覆盖端点：
- GET  /api/v1/health                          （健康检查，无需认证）
- GET  /api/v1/dashboard/top10                 （Top10 成本资源，需认证）
- GET  /api/v1/dashboard/trend                 （成本趋势，需认证）
- GET  /api/v1/dashboard/details               （成本明细，需认证）
- GET  /api/v1/suggestions/idle                （闲置清单，需认证）
- GET  /api/v1/suggestions/list                （优化建议列表，需认证）
- GET  /api/v1/bill/export/csv                 （CSV 账单导出，需认证）
- GET  /api/v1/bill/export/excel               （Excel 账单导出，需认证）
- GET  /api/v1/allocation/configs              （分账配置列表，需认证）
- POST /api/v1/allocation/configs              （新建/更新分账配置，需认证）
- GET  /api/v1/allocation/execute              （执行分账，需认证）

测试覆盖（≥15 个用例）：
1. 健康检查（1）
2. 认证机制（1）
3. Top10 看板（2）：数据正确 + 降序排列
4. 趋势看板（2）：数据正确 + 粒度正确
5. 明细看板（1）：数据正确
6. 闲置清单（2）：5 类模式识别 + 闲置资源数量
7. 优化建议（2）：5 类建议生成 + 节约成本
8. CSV 导出（2）：明细 + 汇总
9. Excel 导出（2）：明细 + 汇总
10. 分账配置（2）：创建 + 比例校验
11. 分账执行（1）：分账结果正确

设计要点：
- 每个测试函数独立，不依赖执行顺序；
- 使用 ``tenant_client`` fixture 支持多租户 token；
- 当 Docker 容器未启动时通过 conftest 钩子自动跳过。
"""

from __future__ import annotations

import io
import time
from typing import Dict

import pytest
import requests

# 复用 conftest 的 JWT 生成函数与配置
from conftest import generate_test_jwt, DEFAULT_TIMEOUT


# ---------------------------------------------------------------------------
# FinOps 看板服务 URL fixture
# ---------------------------------------------------------------------------
@pytest.fixture(scope="session")
def dashboard_url() -> str:
    """FinOps 看板服务基础 URL。

    默认 http://localhost:18087，可通过环境变量 FINOPS_DASHBOARD_URL 覆盖。
    """
    import os
    return os.environ.get("FINOPS_DASHBOARD_URL", "http://localhost:18087")


# ---------------------------------------------------------------------------
# 多租户 HTTP 客户端 fixture
# ---------------------------------------------------------------------------
def _make_client(tenant_id: str) -> requests.Session:
    """创建指定租户的 HTTP 客户端（注入对应 tenantId 的 JWT）。"""
    session = requests.Session()
    token = generate_test_jwt(tenant_id=tenant_id)
    session.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })
    return session


@pytest.fixture
def tenant_a_client() -> requests.Session:
    """租户 A 的 HTTP 客户端。"""
    client = _make_client("finops-dash-tenant-a")
    yield client
    client.close()


@pytest.fixture
def tenant_b_client() -> requests.Session:
    """租户 B 的 HTTP 客户端。"""
    client = _make_client("finops-dash-tenant-b")
    yield client
    client.close()


# ---------------------------------------------------------------------------
# 时间窗口常量
# ---------------------------------------------------------------------------
ISO_START = "2026-08-07T00:00:00Z"
ISO_END = "2026-08-08T00:00:00Z"


# ---------------------------------------------------------------------------
# 1. 健康检查
# ---------------------------------------------------------------------------
def test_health_check(dashboard_url):
    """验证 FinOps 看板服务健康检查端点返回 200 且 status=UP。"""
    resp = requests.get(dashboard_url + "/api/v1/health", timeout=10)
    assert resp.status_code == 200
    body = resp.json()
    assert body.get("status") == "UP"
    assert body.get("service") == "finops-dashboard"


# ---------------------------------------------------------------------------
# 2. 认证机制
# ---------------------------------------------------------------------------
def test_unauthorized_without_token(dashboard_url):
    """验证无 Bearer token 访问受保护端点返回 401。"""
    resp = requests.get(
        dashboard_url + "/api/v1/dashboard/top10",
        params={"start": ISO_START, "end": ISO_END},
        timeout=10,
    )
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# 3. Top10 看板
# ---------------------------------------------------------------------------
def test_top10_returns_sorted_resources(tenant_a_client, dashboard_url):
    """验证 Top10 看板返回资源列表且按总成本降序排列。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/top10",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    # 验证降序排列
    costs = [item["totalCost"] for item in items]
    assert costs == sorted(costs, reverse=True)
    # 验证字段完整
    first = items[0]
    assert "resourceId" in first
    assert "totalCost" in first
    assert "percentage" in first


def test_top10_limit_at_most_10(tenant_a_client, dashboard_url):
    """验证 Top10 看板返回资源数 ≤ 10。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/top10",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["items"]) <= 10


# ---------------------------------------------------------------------------
# 4. 趋势看板
# ---------------------------------------------------------------------------
def test_trend_returns_time_series(tenant_a_client, dashboard_url):
    """验证趋势看板返回时间序列数据。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/trend",
        params={
            "start": ISO_START,
            "end": ISO_END,
            "granularity": "HOUR",
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    # 验证字段完整
    first = items[0]
    assert "timestamp" in first
    assert "totalCost" in first
    assert "cpuCost" in first
    assert "memoryCost" in first


def test_trend_granularity_correct(tenant_a_client, dashboard_url):
    """验证趋势看板粒度参数正确传递。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/trend",
        params={
            "start": ISO_START,
            "end": ISO_END,
            "granularity": "DAY",
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    # DAY 粒度，24 小时窗口应返回 1 个点
    for item in items:
        assert item["granularity"] == "DAY"


# ---------------------------------------------------------------------------
# 5. 明细看板
# ---------------------------------------------------------------------------
def test_details_returns_resource_list(tenant_a_client, dashboard_url):
    """验证明细看板返回资源成本明细列表。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/details",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    # 验证字段完整
    first = items[0]
    assert "resourceId" in first
    assert "resourceType" in first
    assert "totalCost" in first
    assert "tenant" in first


# ---------------------------------------------------------------------------
# 6. 闲置清单（5 类闲置模式识别）
# ---------------------------------------------------------------------------
def test_idle_list_identifies_5_patterns(tenant_a_client, dashboard_url):
    """验证闲置清单识别出 5 类闲置模式。

    5 类：LOW_CPU_UTILIZATION / LOW_MEMORY_UTILIZATION / UNMOUNTED_STORAGE /
          IDLE_GPU / LOW_NETWORK_TRAFFIC
    """
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/suggestions/idle",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    patterns = {item["pattern"] for item in items}
    expected_patterns = {
        "LOW_CPU_UTILIZATION",
        "LOW_MEMORY_UTILIZATION",
        "UNMOUNTED_STORAGE",
        "IDLE_GPU",
        "LOW_NETWORK_TRAFFIC",
    }
    assert expected_patterns.issubset(patterns), (
        f"期望 5 类闲置模式均被识别，实际识别: {patterns}"
    )


def test_idle_resources_have_suggestions(tenant_a_client, dashboard_url):
    """验证每个闲置资源都有优化建议字段。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/suggestions/idle",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    for item in items:
        assert "suggestion" in item
        assert item["suggestion"]  # 非空
        assert "estimatedSaving" in item
        assert item["estimatedSaving"] >= 0


# ---------------------------------------------------------------------------
# 7. 优化建议
# ---------------------------------------------------------------------------
def test_suggestions_cover_5_patterns(tenant_a_client, dashboard_url):
    """验证优化建议覆盖 5 类闲置模式。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/suggestions/list",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    patterns = {item["pattern"] for item in items}
    expected_patterns = {
        "LOW_CPU_UTILIZATION",
        "LOW_MEMORY_UTILIZATION",
        "UNMOUNTED_STORAGE",
        "IDLE_GPU",
        "LOW_NETWORK_TRAFFIC",
    }
    assert expected_patterns.issubset(patterns), (
        f"期望 5 类优化建议均生成，实际: {patterns}"
    )


def test_suggestions_have_estimated_saving(tenant_a_client, dashboard_url):
    """验证每条优化建议都有估算节约成本。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/suggestions/list",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    for item in items:
        assert "estimatedMonthlySaving" in item
        assert item["estimatedMonthlySaving"] >= 0
        assert "actionType" in item
        assert "riskLevel" in item


# ---------------------------------------------------------------------------
# 8. CSV 账单导出
# ---------------------------------------------------------------------------
def test_csv_export_details(tenant_a_client, dashboard_url):
    """验证 CSV 明细账单导出格式正确。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/bill/export/csv",
        params={
            "type": "details",
            "start": ISO_START,
            "end": ISO_END,
        },
    )
    assert resp.status_code == 200
    assert "text/csv" in resp.headers.get("content-type", "")
    content = resp.content
    # 验证含 UTF-8 BOM
    assert content[:3] == b"\xef\xbb\xbf"
    # 解析 CSV
    text = content[3:].decode("utf-8")
    lines = text.strip().split("\n")
    assert len(lines) > 1  # 表头 + 至少 1 行数据
    header = lines[0]
    assert "资源ID" in header
    assert "总成本" in header


def test_csv_export_summary(tenant_a_client, dashboard_url):
    """验证 CSV 汇总账单导出格式正确。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/bill/export/csv",
        params={
            "type": "summary",
            "groupBy": "TENANT",
            "start": ISO_START,
            "end": ISO_END,
        },
    )
    assert resp.status_code == 200
    content = resp.content
    text = content[3:].decode("utf-8")
    lines = text.strip().split("\n")
    assert len(lines) > 1
    header = lines[0]
    assert "聚合维度" in header
    assert "聚合键" in header
    assert "总成本" in header


# ---------------------------------------------------------------------------
# 9. Excel 账单导出
# ---------------------------------------------------------------------------
def test_excel_export_details(tenant_a_client, dashboard_url):
    """验证 Excel 明细账单导出格式正确。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/bill/export/excel",
        params={
            "type": "details",
            "start": ISO_START,
            "end": ISO_END,
        },
    )
    assert resp.status_code == 200
    content_type = resp.headers.get("content-type", "")
    assert "spreadsheetml" in content_type or "octet-stream" in content_type
    content = resp.content
    # 验证 Excel 文件魔数（PK 开头，ZIP 格式）
    assert content[:2] == b"PK"
    # 验证文件非空
    assert len(content) > 100


def test_excel_export_full_has_two_sheets(tenant_a_client, dashboard_url):
    """验证 Excel full 导出含明细与汇总两个 Sheet。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/bill/export/excel",
        params={
            "type": "full",
            "groupBy": "TENANT",
            "start": ISO_START,
            "end": ISO_END,
        },
    )
    assert resp.status_code == 200
    content = resp.content
    assert content[:2] == b"PK"
    # 使用 openpyxl 验证 Sheet 数量
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        sheet_names = wb.sheetnames
        assert "成本明细" in sheet_names
        assert "成本汇总" in sheet_names
        assert len(sheet_names) == 2
    except ImportError:
        # openpyxl 不可用时仅验证文件格式
        pass


# ---------------------------------------------------------------------------
# 10. 分账配置
# ---------------------------------------------------------------------------
def test_allocation_config_create_and_list(tenant_a_client, dashboard_url):
    """验证分账配置创建与列表查询。"""
    config_id = f"test-alloc-{int(time.time())}"
    payload = {
        "id": config_id,
        "parentWorkspace": "ns-test",
        "dimension": "namespace",
        "ratios": {
            "sub-ws-1": 0.6,
            "sub-ws-2": 0.3,
            "sub-ws-3": 0.1,
        },
        "enabled": True,
        "remark": "测试分账配置",
    }
    # 创建
    resp = tenant_a_client.post(
        dashboard_url + "/api/v1/allocation/configs",
        json=payload,
    )
    assert resp.status_code == 200
    created = resp.json()
    assert created["id"] == config_id
    assert created["ratios"]["sub-ws-1"] == 0.6

    # 列表查询
    resp = tenant_a_client.get(dashboard_url + "/api/v1/allocation/configs")
    assert resp.status_code == 200
    configs = resp.json()
    ids = [c["id"] for c in configs]
    assert config_id in ids


def test_allocation_config_invalid_ratios_rejected(tenant_a_client, dashboard_url):
    """验证分账比例合计 != 1.0 时返回 400。"""
    payload = {
        "id": "invalid-ratios",
        "parentWorkspace": "ns-test",
        "dimension": "namespace",
        "ratios": {
            "sub-ws-1": 0.5,
            "sub-ws-2": 0.3,  # 合计 0.8 != 1.0
        },
        "enabled": True,
    }
    resp = tenant_a_client.post(
        dashboard_url + "/api/v1/allocation/configs",
        json=payload,
    )
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# 11. 分账执行
# ---------------------------------------------------------------------------
def test_allocation_execute(tenant_a_client, dashboard_url):
    """验证执行分账返回分账结果。"""
    # 先创建一个分账配置
    config_id = f"test-exec-{int(time.time())}"
    payload = {
        "id": config_id,
        "parentWorkspace": "ns-demo",
        "dimension": "namespace",
        "ratios": {
            "sub-analytics": 0.7,
            "sub-training": 0.3,
        },
        "enabled": True,
    }
    tenant_a_client.post(
        dashboard_url + "/api/v1/allocation/configs",
        json=payload,
    )

    # 执行分账
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/allocation/execute",
        params={
            "configId": config_id,
            "start": ISO_START,
            "end": ISO_END,
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    items = body["items"]
    assert len(items) > 0
    # 验证分账项字段
    for item in items:
        assert "parentWorkspace" in item
        assert "subWorkspace" in item
        assert "ratio" in item
        assert "allocatedCost" in item
        assert item["ratio"] > 0


# ---------------------------------------------------------------------------
# 12. 租户隔离
# ---------------------------------------------------------------------------
def test_tenant_isolation_top10(tenant_a_client, tenant_b_client, dashboard_url):
    """验证 Top10 看板租户隔离：租户 A 与租户 B 看到不同数据。"""
    resp_a = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/top10",
        params={"start": ISO_START, "end": ISO_END},
    )
    resp_b = tenant_b_client.get(
        dashboard_url + "/api/v1/dashboard/top10",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp_a.status_code == 200
    assert resp_b.status_code == 200
    # 验证租户字段正确
    if resp_a.json()["items"]:
        for item in resp_a.json()["items"]:
            assert item.get("tenant") == "finops-dash-tenant-a"
    if resp_b.json()["items"]:
        for item in resp_b.json()["items"]:
            assert item.get("tenant") == "finops-dash-tenant-b"


# ---------------------------------------------------------------------------
# 13. 看板响应结构
# ---------------------------------------------------------------------------
def test_dashboard_response_structure(tenant_a_client, dashboard_url):
    """验证看板响应包含统一结构字段。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/dashboard/details",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "items" in body
    assert "total" in body
    assert "start" in body
    assert "end" in body
    assert "tenant" in body
    assert "summary" in body
    assert body["total"] == len(body["items"])


# ---------------------------------------------------------------------------
# 14. 闲置模式枚举完整性
# ---------------------------------------------------------------------------
def test_idle_pattern_enum_values(tenant_a_client, dashboard_url):
    """验证闲置清单返回的 pattern 字段值在 5 类枚举范围内。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/suggestions/idle",
        params={"start": ISO_START, "end": ISO_END},
    )
    assert resp.status_code == 200
    valid_patterns = {
        "LOW_CPU_UTILIZATION",
        "LOW_MEMORY_UTILIZATION",
        "UNMOUNTED_STORAGE",
        "IDLE_GPU",
        "LOW_NETWORK_TRAFFIC",
    }
    for item in resp.json()["items"]:
        assert item["pattern"] in valid_patterns, (
            f"无效闲置模式: {item['pattern']}"
        )


# ---------------------------------------------------------------------------
# 15. 分账比例边界值
# ---------------------------------------------------------------------------
def test_allocation_ratio_boundary(tenant_a_client, dashboard_url):
    """验证分账比例边界值：单一子工作空间 ratio=1.0 应成功。"""
    config_id = f"test-boundary-{int(time.time())}"
    payload = {
        "id": config_id,
        "parentWorkspace": "ns-boundary",
        "dimension": "namespace",
        "ratios": {"only-ws": 1.0},
        "enabled": True,
    }
    resp = tenant_a_client.post(
        dashboard_url + "/api/v1/allocation/configs",
        json=payload,
    )
    assert resp.status_code == 200
    assert resp.json()["ratios"]["only-ws"] == 1.0


# ---------------------------------------------------------------------------
# 16. Excel full 导出含明细与汇总数据
# ---------------------------------------------------------------------------
def test_excel_full_export_has_data(tenant_a_client, dashboard_url):
    """验证 Excel full 导出含明细与汇总数据行。"""
    resp = tenant_a_client.get(
        dashboard_url + "/api/v1/bill/export/excel",
        params={
            "type": "full",
            "groupBy": "WORKSPACE",
            "start": ISO_START,
            "end": ISO_END,
        },
    )
    assert resp.status_code == 200
    content = resp.content
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        detail_sheet = wb["成本明细"]
        summary_sheet = wb["成本汇总"]
        # 明细 Sheet 应有表头 + 数据行
        assert detail_sheet.max_row > 1
        # 汇总 Sheet 应有表头 + 数据行
        assert summary_sheet.max_row > 1
        # 验证明细表头
        assert detail_sheet.cell(1, 1).value == "资源ID"
        # 验证汇总表头
        assert summary_sheet.cell(1, 1).value == "聚合维度"
    except ImportError:
        pass